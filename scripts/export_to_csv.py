#!/usr/bin/env python3
"""
Deflation Index - CSV Export Script v3.0.1
Exports data from Excel files to CSV format
"""

from openpyxl import load_workbook
import csv
from pathlib import Path

def export_master_to_csv():
    """Export Master_Index sheet to CSV"""
    
    excel_path = Path('data/excel/master_deflation_index_v3.0.1.xlsx')
    csv_path = Path('data/csv/master_di_1990_2024.csv')
    
    # Create csv directory if it doesn't exist
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    
    print(f"Loading {excel_path}...")
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Master_Index']
    
    # Find data range
    headers = []
    for col in range(1, 13):  # Columns A-L
        header = ws.cell(7, col).value
        if header:
            headers.append(header)
    
    # Export data
    with open(csv_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        
        # Rows 8-42 contain data (1990-2024)
        for row in range(8, 43):
            row_data = []
            for col in range(1, len(headers) + 1):
                value = ws.cell(row, col).value
                row_data.append(value if value is not None else '')
            writer.writerow(row_data)
    
    wb.close()
    print(f"✓ Exported to {csv_path}")
    return csv_path

def export_sector_to_csv(sector_name, excel_filename, start_year, end_year):
    """Export sector index data to CSV"""
    
    excel_path = Path(f'data/excel/{excel_filename}')
    csv_path = Path(f'data/csv/{sector_name.lower()}_{start_year}_{end_year}.csv')
    
    # Create csv directory if it doesn't exist
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    
    if not excel_path.exists():
        print(f"⚠ Skipping {sector_name}: {excel_path} not found")
        return None
    
    print(f"Loading {excel_path}...")
    wb = load_workbook(excel_path, data_only=True)
    
    # Try to find Sector_Index sheet
    if 'Sector_Index' not in wb.sheetnames:
        print(f"⚠ Skipping {sector_name}: No Sector_Index sheet")
        wb.close()
        return None
    
    ws = wb['Sector_Index']
    
    # Export data (adjust based on actual sheet structure)
    with open(csv_path, 'w', newline='') as f:
        writer = csv.writer(f)
        
        # Write headers
        headers = ['Year', f'{sector_name}_Index']
        writer.writerow(headers)
        
        # Write data - this is a simplified version
        # Actual implementation would need to match your Excel structure
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If year exists
                writer.writerow([row[0], row[1] if len(row) > 1 else ''])
    
    wb.close()
    print(f"✓ Exported to {csv_path}")
    return csv_path

def main():
    """Export all data to CSV"""
    
    print("="*80)
    print("DEFLATION INDEX - CSV EXPORT v3.0.1")
    print("="*80)
    print()
    
    # Export master index
    export_master_to_csv()
    print()
    
    # Export sector indices
    sectors = [
        ('Computing', 'computing_deflation_index_v1.0.xlsx', 1990, 2024),
        ('Communications', 'communications_deflation_index_v1.0.xlsx', 1990, 2024),
        ('Energy', 'energy_deflation_index_v1.0.xlsx', 1990, 2024),
        ('Transportation', 'transportation_deflation_index_v1.0.xlsx', 2010, 2024),
    ]
    
    for sector_name, filename, start_year, end_year in sectors:
        export_sector_to_csv(sector_name, filename, start_year, end_year)
    
    print()
    print("="*80)
    print("CSV EXPORT COMPLETE")
    print("="*80)
    print()
    print("CSV files created in data/csv/")
    print()
    print("Note: Sector CSV exports are simplified.")
    print("Review and adjust based on your Excel structure.")

if __name__ == "__main__":
    main()
